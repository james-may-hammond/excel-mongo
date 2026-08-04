import asyncio
import base64
import json
from datetime import datetime

import httpx
from fastapi import APIRouter, WebSocket, WebSocketDisconnect

from be.db import get_ws_db, FINNOTO_COMMON_HEADERS
from be.routes.fetch import FetchRequest as FetchReq, fetch
from be.routes.schema import get_schema
from be.routes.reports import (
    ReportSheetRequest,
    DateFilterDto,
    DateRangeDto,
    _post as reports_post,
    _get as reports_get,
    _normalise_sheet_response,
)


def _json_serializer(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()
    if isinstance(obj, bytes):
        return base64.b64encode(obj).decode('utf-8')
    raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable")


router = APIRouter()


class ConnectionManager:
    def __init__(self):
        self.active_connections: list[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def send_response(self, websocket: WebSocket, request_id: str, status: str, data: dict = None, error: str = None):
        response = {"requestID": request_id, "status": status}
        if data is not None:
            response["data"] = data
        if error is not None:
            response["error"] = error
        await websocket.send_text(json.dumps(response))


manager = ConnectionManager()


async def _stream_sheets_via_json(websocket, request_id, db, payload, base_url=None):
    """
    Primary path: fetch sheet data via get-report-sheet JSON API.
    Response shape: { "SheetName": [ {col: val, ...}, ... ], ... }
    """
    base_url = base_url or db.base_url
    body = {
        "business_report_id": payload["business_report_id"],
        "sheet_ids": payload["sheet_ids"],
    }
    if payload.get("date"):
        body["date"] = payload["date"]
    if payload.get("filters"):
        body["filters"] = payload["filters"]

    await websocket.send_text(json.dumps({
        "requestId": request_id,
        "status": "progress",
        "data": {"message": "Fetching sheet data from Finnoto..."}
    }))

    raw = await reports_post(
        f"{base_url}/api/b/report/get-report-sheet",
        body,
        db.token
    )

    sheets = _normalise_sheet_response(raw)
    if not sheets:
        await manager.send_response(websocket, request_id, "error",
                                    error="No sheet data returned from get-report-sheet.")
        return

    total_rows = 0
    for sheet in sheets:
        cols = sheet.get("columns", [])
        rows = sheet.get("rows", [])
        sheet_name = sheet.get("name", "Sheet")

        await websocket.send_text(json.dumps({
            "requestId": request_id,
            "status": "sheet_start",
            "data": {
                "_sheet_name": sheet_name,
                "_columns": cols,
                "_row_count": len(rows),
            }
        }, default=_json_serializer))

        BATCH = 500
        for i in range(0, len(rows), BATCH):
            batch = rows[i:i + BATCH]
            await websocket.send_text(json.dumps({
                "requestId": request_id,
                "status": "chunk",
                "data": {
                    "_sheet_name": sheet_name,
                    "_rows": batch,
                }
            }, default=_json_serializer))
            await asyncio.sleep(0)

        await websocket.send_text(json.dumps({
            "requestId": request_id,
            "status": "sheet_complete",
            "data": {"_sheet_name": sheet_name}
        }))
        total_rows += len(rows)

    await manager.send_response(websocket, request_id, "success",
                                data={"message": f"Report imported: {total_rows} rows across {len(sheets)} sheet(s)"})


async def _stream_sheets_via_generate(websocket, request_id, db, payload, base_url=None):
    """
    Fallback path: generate report, download .xlsx, parse with openpyxl.
    Kept from the original implementation.
    """
    base_url = base_url or db.base_url
    report_url = payload.get("url")
    business_report_id = payload.get("business_report_id")
    date_raw = payload.get("date")

    if not report_url:
        await websocket.send_text(json.dumps({
            "requestId": request_id,
            "status": "progress",
            "data": {"message": "Generating report on Finnoto..."}
        }))

        params = {}
        if date_raw:
            params["date_filter"] = date_raw
        gen_body = {"params": params}

        try:
            new_report = await reports_post(
                f"{base_url}/api/b/report/{business_report_id}/generate-report",
                gen_body,
                db.token
            )
        except Exception as e:
            await manager.send_response(websocket, request_id, "error",
                                        error=f"Failed to generate report: {getattr(e, 'detail', str(e))}")
            return

        report_id = new_report.get("id")
        if not report_id:
            await manager.send_response(websocket, request_id, "error",
                                        error="Finnoto did not return a report ID after generation.")
            return

        await websocket.send_text(json.dumps({
            "requestId": request_id,
            "status": "progress",
            "data": {"message": f"Report {report_id} queued. Waiting for processing..."}
        }))

        for attempt in range(30):
            await asyncio.sleep(2)
            try:
                async with httpx.AsyncClient(timeout=15.0) as client:
                    poll_res = await client.get(
                        f"{base_url}/api/b/report/{report_id}",
                        headers={**FINNOTO_COMMON_HEADERS, "Authorization": f"Bearer {db.token}"}
                    )
                if poll_res.status_code == 200:
                    poll_data = poll_res.json()
                    if poll_data.get("url") or poll_data.get("processed_at"):
                        report_url = poll_data.get("url")
                        break
            except Exception:
                pass

            await websocket.send_text(json.dumps({
                "requestId": request_id,
                "status": "progress",
                "data": {"message": f"Processing... ({attempt + 1}/30)"}
            }))

        if not report_url:
            await manager.send_response(websocket, request_id, "error",
                                        error="Report generation timed out. Please try again later.")
            return

    await websocket.send_text(json.dumps({
        "requestId": request_id,
        "status": "progress",
        "data": {"message": "Downloading report..."}
    }))

    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            dl_res = await client.get(report_url)

        if dl_res.status_code == 200:
            content_type = dl_res.headers.get("content-type", "")
            if "spreadsheet" in content_type or "excel" in content_type or report_url.endswith(".xlsx") or len(dl_res.content) > 100:
                import io
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(io.BytesIO(dl_res.content), read_only=True, data_only=True)
                    for sheet_name in wb.sheetnames:
                        ws_sheet = wb[sheet_name]
                        rows = list(ws_sheet.values())
                        if not rows:
                            continue
                        columns = [str(c) if c is not None else f"Col {i+1}" for i, c in enumerate(rows[0])]
                        data_rows = [list(row) for row in rows[1:]]

                        await websocket.send_text(json.dumps({
                            "requestId": request_id,
                            "status": "sheet_start",
                            "data": {
                                "_sheet_name": sheet_name,
                                "_columns": columns,
                                "_row_count": len(data_rows),
                            }
                        }, default=_json_serializer))

                        BATCH = 500
                        for i in range(0, len(data_rows), BATCH):
                            batch = data_rows[i:i + BATCH]
                            await websocket.send_text(json.dumps({
                                "requestId": request_id,
                                "status": "chunk",
                                "data": {
                                    "_sheet_name": sheet_name,
                                    "_rows": batch,
                                }
                            }, default=_json_serializer))
                            await asyncio.sleep(0)

                        await websocket.send_text(json.dumps({
                            "requestId": request_id,
                            "status": "sheet_complete",
                            "data": {"_sheet_name": sheet_name}
                        }))

                    wb.close()
                    await manager.send_response(websocket, request_id, "success",
                                                data={"message": "Report imported successfully", "url": report_url})
                except ImportError:
                    await manager.send_response(websocket, request_id, "success",
                                                data={"message": "Report downloaded. openpyxl not installed for .xlsx parsing.",
                                                       "url": report_url})
            else:
                await manager.send_response(websocket, request_id, "success",
                                            data={"message": "Downloaded but not in .xlsx format.",
                                                   "url": report_url, "content_type": content_type})
        else:
            await manager.send_response(websocket, request_id, "success",
                                        data={"message": f"Download failed (HTTP {dl_res.status_code}). Open the URL manually.",
                                               "url": report_url})
    except Exception as e:
        await manager.send_response(websocket, request_id, "success",
                                    data={"message": "Could not auto-download. Open the URL manually.",
                                           "url": report_url, "error": str(e)})


@router.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket, token: str, base_url: str = None):
    await manager.connect(websocket)
    try:
        db = await get_ws_db(token, base_url or "")
    except Exception:
        await manager.send_response(websocket, "auth", "error", error="Invalid token or connection failed")
        await websocket.close()
        return

    try:
        while True:
            text = await websocket.receive_text()
            try:
                msg = json.loads(text)
            except json.JSONDecodeError:
                continue

            request_id = msg.get("requestId")
            action = msg.get("action")
            payload = msg.get("payload", {})
            if not request_id or not action:
                continue

            try:
                if action == "stream_fetch":
                    col = db[payload.get("collection")]
                    cursor = col.find(payload.get("filters", {}))
                    if payload.get("limit", 0) > 0:
                        cursor = cursor.limit(payload["limit"])

                    count = 0
                    async for doc in cursor.batch_size(1000):
                        await websocket.send_text(json.dumps({
                            "requestId": request_id,
                            "status": "chunk",
                            "data": doc
                        }, default=_json_serializer))
                        count += 1
                        if count % 2000 == 0:
                            await asyncio.sleep(0.5)
                    await manager.send_response(websocket, request_id, "success",
                                                data={"message": "Stream complete", "total": count})

                elif action == "multi_stream_fetch":
                    queries = payload.get("queries", [])
                    for query in queries:
                        col = db[query.get("collection")]
                        cursor = col.find(query.get("filters", {}))
                        if query.get("limit", 0) > 0:
                            cursor = cursor.limit(query["limit"])
                        async for doc in cursor.batch_size(1000):
                            await websocket.send_text(json.dumps({
                                "requestId": request_id,
                                "status": "chunk",
                                "data": {**doc, "_target_collection": query["collection"]}
                            }, default=_json_serializer))
                    await manager.send_response(websocket, request_id, "success",
                                                data={"message": "Multi-stream complete"})

                elif action == "report_sheet_stream":
                    """
                    Stream business report sheet data into Excel.

                    Primary path: uses get-report-sheet JSON API (fast, no generation).
                    Fallback path: generate → download .xlsx → parse openpyxl.
                    """
                    fallback = payload.get("_fallback", False)
                    business_report_id = payload.get("business_report_id")
                    sheet_ids = payload.get("sheet_ids")
                    report_base = payload.get("base_url") or db.base_url

                    if not business_report_id:
                        await manager.send_response(websocket, request_id, "error",
                                                    error="business_report_id is required")
                        continue

                    # sheet_ids are required only for the JSON API path; fallback uses URL/generate
                    if not fallback and not sheet_ids:
                        await manager.send_response(websocket, request_id, "error",
                                                    error="sheet_ids are required when not using fallback mode")
                        continue

                    if not fallback:
                        try:
                            await _stream_sheets_via_json(websocket, request_id, db, payload, base_url=report_base)
                            continue
                        except Exception as e:
                            err_detail = getattr(e, "detail", str(e))
                            await websocket.send_text(json.dumps({
                                "requestId": request_id,
                                "status": "progress",
                                "data": {"message": f"JSON API failed ({err_detail}). Falling back to generate flow..."}
                            }))

                    await _stream_sheets_via_generate(websocket, request_id, db, payload, base_url=report_base)

                elif action == "fetch":
                    res = await fetch(request=FetchReq(**payload), db=db)
                    await manager.send_response(websocket, request_id, "success", data=res)

                elif action == "schema":
                    res = await get_schema(collection=payload.get("collection"), db=db)
                    await manager.send_response(websocket, request_id, "success", data=res)

                else:
                    await manager.send_response(websocket, request_id, "error",
                                                error=f"Unknown action: {action}")

            except Exception as e:
                await manager.send_response(websocket, request_id, "error",
                                            error=getattr(e, "detail", str(e)))

    except WebSocketDisconnect:
        manager.disconnect(websocket)
